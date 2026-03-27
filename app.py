from flask import Flask, render_template, request, redirect, url_for, session, send_file
import os
import io
import sqlite3
import qrcode
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
app.secret_key = "pizarro123"

# =========================
# CONFIG
# =========================
IS_RAILWAY = os.environ.get("RAILWAY_ENVIRONMENT") is not None
DATA_DIR   = "/data" if IS_RAILWAY else os.path.join(os.path.dirname(__file__), "data")
DB_PATH    = os.path.join(DATA_DIR, "pizarro.db")
QR_FOLDER  = "static/qrcodes"

BASE_URL = os.environ.get(
    "BASE_URL",
    "https://movelaria-pizarro-production.up.railway.app"
).rstrip("/")

os.makedirs(DATA_DIR,  exist_ok=True)
os.makedirs(QR_FOLDER, exist_ok=True)

# =========================
# BANCO DE DADOS
# =========================
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with get_db() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS sobras (
                id        INTEGER PRIMARY KEY AUTOINCREMENT,
                largura   TEXT NOT NULL,
                altura    TEXT NOT NULL,
                espessura TEXT,
                cor       TEXT,
                obs       TEXT,
                usado     TEXT NOT NULL DEFAULT 'NÃO'
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS usuarios (
                id    INTEGER PRIMARY KEY AUTOINCREMENT,
                user  TEXT UNIQUE NOT NULL,
                senha TEXT NOT NULL,
                role  TEXT NOT NULL DEFAULT 'funcionario'
            )
        """)
        # Migração: adiciona coluna role se não existir (banco antigo)
        try:
            conn.execute("ALTER TABLE usuarios ADD COLUMN role TEXT NOT NULL DEFAULT 'funcionario'")
            conn.commit()
        except Exception:
            pass

        # Migração: primeiro usuário cadastrado vira admin se nenhum admin existir
        try:
            admins = conn.execute("SELECT COUNT(*) FROM usuarios WHERE role = 'admin'").fetchone()[0]
            if admins == 0:
                primeiro = conn.execute("SELECT id FROM usuarios ORDER BY id LIMIT 1").fetchone()
                if primeiro:
                    conn.execute("UPDATE usuarios SET role = 'admin' WHERE id = ?", (primeiro[0],))
                    conn.commit()
        except Exception:
            pass

        conn.commit()

init_db()

# =========================
# HELPERS
# =========================
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def is_admin():
    return session.get("role") == "admin"

def requer_login(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

def requer_admin(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session:
            return redirect(url_for("login"))
        if not is_admin():
            return redirect(url_for("index"))
        return f(*args, **kwargs)
    return decorated

def gerar_qr(item_id):
    url = f"{BASE_URL}/sobra/{item_id}"
    qr = qrcode.make(url)
    qr.save(f"{QR_FOLDER}/qr_{item_id}.png")

def regenerar_todos_qrs():
    try:
        with get_db() as conn:
            rows = conn.execute("SELECT id FROM sobras").fetchall()
        for row in rows:
            if not os.path.exists(f"{QR_FOLDER}/qr_{row['id']}.png"):
                gerar_qr(row["id"])
    except Exception:
        pass

regenerar_todos_qrs()

# =========================
# LOGIN
# =========================
@app.route("/login", methods=["GET", "POST"])
def login():
    if "user" in session:
        return redirect(url_for("index"))

    erro = None
    if request.method == "POST":
        user  = request.form.get("user",  "").strip()
        senha = request.form.get("senha", "").strip()

        with get_db() as conn:
            row = conn.execute(
                "SELECT * FROM usuarios WHERE user = ? AND senha = ?",
                (user, senha)
            ).fetchone()

        if row:
            session["user"] = user
            session["role"] = row["role"]
            return redirect(url_for("index"))
        else:
            erro = "Usuário ou senha incorretos."

    return render_template("login.html", erro=erro)

# =========================
# CADASTRO
# =========================
@app.route("/register", methods=["GET", "POST"])
def register():
    erro = None
    if request.method == "POST":
        user   = request.form.get("user",   "").strip()
        senha  = request.form.get("senha",  "").strip()
        senha2 = request.form.get("senha2", "").strip()

        if not user or not senha:
            erro = "Preencha todos os campos."
        elif senha != senha2:
            erro = "As senhas não coincidem."
        else:
            try:
                with get_db() as conn:
                    # Primeiro usuário vira admin automaticamente
                    total = conn.execute("SELECT COUNT(*) FROM usuarios").fetchone()[0]
                    role  = "admin" if total == 0 else "funcionario"
                    conn.execute(
                        "INSERT INTO usuarios (user, senha, role) VALUES (?, ?, ?)",
                        (user, senha, role)
                    )
                    conn.commit()
                return redirect(url_for("login"))
            except sqlite3.IntegrityError:
                erro = "Este usuário já existe. Escolha outro."

    return render_template("register.html", erro=erro)

# =========================
# HOME — ESTOQUE
# =========================
@app.route("/", methods=["GET", "POST"])
@requer_login
def index():
    if request.method == "POST" and is_admin():
        largura   = request.form.get("largura",   "").strip()
        altura    = request.form.get("altura",    "").strip()
        espessura = request.form.get("espessura", "").strip()
        cor       = request.form.get("cor",       "").strip()
        obs       = request.form.get("obs",       "").strip()

        if largura and altura:
            with get_db() as conn:
                cursor = conn.execute(
                    "INSERT INTO sobras (largura, altura, espessura, cor, obs, usado) VALUES (?, ?, ?, ?, ?, 'NÃO')",
                    (largura, altura, espessura, cor, obs)
                )
                conn.commit()
                gerar_qr(cursor.lastrowid)

        return redirect(url_for("index"))

    with get_db() as conn:
        rows = conn.execute("SELECT * FROM sobras ORDER BY id DESC").fetchall()

    dados = [dict(row) for row in rows]
    return render_template("index.html", dados=dados, user=session["user"], admin=is_admin())

# =========================
# DETALHE (via QR Code)
# =========================
@app.route("/sobra/<int:id>")
def detalhe(id):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM sobras WHERE id = ?", (id,)).fetchone()

    if not row:
        return "Peça não encontrada.", 404

    return render_template("detalhe.html", sobra=dict(row))

# =========================
# USAR PEÇA
# =========================
@app.route("/usar/<int:id>")
@requer_login
def usar(id):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM sobras WHERE id = ?", (id,)).fetchone()
        if not row:
            return "Peça não encontrada.", 404
        if row["usado"] == "NÃO":
            conn.execute("UPDATE sobras SET usado = ? WHERE id = ?", (session["user"], id))
            conn.commit()

    return redirect(url_for("index"))

# =========================
# REATIVAR PEÇA (admin)
# =========================
@app.route("/reativar/<int:id>")
@requer_admin
def reativar(id):
    with get_db() as conn:
        conn.execute("UPDATE sobras SET usado = 'NÃO' WHERE id = ?", (id,))
        conn.commit()
    return redirect(url_for("index"))

# =========================
# DELETAR PEÇA (admin)
# =========================
@app.route("/deletar/<int:id>")
@requer_admin
def deletar(id):
    with get_db() as conn:
        conn.execute("DELETE FROM sobras WHERE id = ?", (id,))
        conn.commit()
    qr_path = f"{QR_FOLDER}/qr_{id}.png"
    if os.path.exists(qr_path):
        os.remove(qr_path)
    return redirect(url_for("index"))

# =========================
# EXPORTAR EXCEL (admin)
# =========================
@app.route("/exportar")
@requer_admin
def exportar():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM sobras ORDER BY id").fetchall()

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Estoque de Sobras"

    hfill = PatternFill("solid", fgColor="2E1F14")
    hfont = Font(bold=True, color="C4864A", size=11)
    halign = Alignment(horizontal="center", vertical="center")

    headers = ["ID", "Largura (mm)", "Altura (mm)", "Espessura", "Cor / Material", "Observação", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill; cell.font = hfont; cell.alignment = halign

    ws.row_dimensions[1].height = 22
    for i, w in enumerate([8,14,14,12,20,25,18], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    fill_d = PatternFill("solid", fgColor="1a3a25")
    fill_u = PatternFill("solid", fgColor="2a1a10")
    font_d = Font(color="6aab7e")
    font_u = Font(color="8a6a5a")

    for r, row in enumerate(rows, 2):
        status = "Disponível" if row["usado"] == "NÃO" else f"Usado por: {row['usado']}"
        values = [row["id"], row["largura"], row["altura"],
                  row["espessura"] or "", row["cor"] or "", row["obs"] or "", status]
        disp = row["usado"] == "NÃO"
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = Alignment(horizontal="center")
            if col == 7:
                cell.fill = fill_d if disp else fill_u
                cell.font = font_d if disp else font_u

    # Aba modelo
    ws2 = wb.create_sheet("Modelo Importação")
    for col, h in enumerate(["largura","altura","espessura","cor","obs"], 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = hfill; cell.font = hfont; cell.alignment = halign
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16
    for col, val in enumerate(["600","1200","15mm","Branco Polar","Sobra lateral"], 1):
        ws2.cell(row=2, column=col, value=val).alignment = Alignment(horizontal="center")
    ws2.cell(row=4, column=1, value="⚠ Preencha a partir da linha 2. Não altere os cabeçalhos.")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name="estoque_pizarro.xlsx")

# =========================
# IMPORTAR EXCEL (admin)
# =========================
@app.route("/importar", methods=["POST"])
@requer_admin
def importar():
    arquivo = request.files.get("planilha")
    if not arquivo or arquivo.filename == "":
        return redirect(url_for("index"))

    try:
        wb = openpyxl.load_workbook(arquivo)
        ws = wb["Modelo Importação"] if "Modelo Importação" in wb.sheetnames else wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0] and not row[1]:
                continue
            try:
                largura   = str(row[0]).strip() if row[0] else ""
                altura    = str(row[1]).strip() if row[1] else ""
                espessura = str(row[2]).strip() if row[2] else ""
                cor       = str(row[3]).strip() if row[3] else ""
                obs       = str(row[4]).strip() if row[4] else ""
                if not largura or not altura:
                    continue
                with get_db() as conn:
                    cursor = conn.execute(
                        "INSERT INTO sobras (largura, altura, espessura, cor, obs, usado) VALUES (?, ?, ?, ?, ?, 'NÃO')",
                        (largura, altura, espessura, cor, obs)
                    )
                    conn.commit()
                    gerar_qr(cursor.lastrowid)
            except Exception:
                continue
    except Exception:
        pass

    return redirect(url_for("index"))

# =========================
# GERENCIAR USUÁRIOS (admin)
# =========================
@app.route("/usuarios")
@requer_admin
def usuarios():
    with get_db() as conn:
        rows = conn.execute("SELECT id, user, role FROM usuarios ORDER BY role, user").fetchall()
    return render_template("usuarios.html", usuarios=[dict(r) for r in rows], user=session["user"])

@app.route("/usuarios/promover/<int:id>")
@requer_admin
def promover(id):
    with get_db() as conn:
        conn.execute("UPDATE usuarios SET role = 'admin' WHERE id = ?", (id,))
        conn.commit()
    return redirect(url_for("usuarios"))

@app.route("/usuarios/rebaixar/<int:id>")
@requer_admin
def rebaixar(id):
    with get_db() as conn:
        # Não pode rebaixar a si mesmo
        row = conn.execute("SELECT user FROM usuarios WHERE id = ?", (id,)).fetchone()
        if row and row["user"] != session["user"]:
            conn.execute("UPDATE usuarios SET role = 'funcionario' WHERE id = ?", (id,))
            conn.commit()
    return redirect(url_for("usuarios"))

@app.route("/usuarios/deletar/<int:id>")
@requer_admin
def deletar_usuario(id):
    with get_db() as conn:
        row = conn.execute("SELECT user FROM usuarios WHERE id = ?", (id,)).fetchone()
        if row and row["user"] != session["user"]:
            conn.execute("DELETE FROM usuarios WHERE id = ?", (id,))
            conn.commit()
    return redirect(url_for("usuarios"))

# =========================
# DASHBOARD (admin)
# =========================
@app.route("/dashboard")
@requer_admin
def dashboard():
    with get_db() as conn:
        total       = conn.execute("SELECT COUNT(*) FROM sobras").fetchone()[0]
        usados      = conn.execute("SELECT COUNT(*) FROM sobras WHERE usado != 'NÃO'").fetchone()[0]
        disponiveis = total - usados
        usados_rows = conn.execute("SELECT largura, altura FROM sobras WHERE usado != 'NÃO'").fetchall()
        ranking_rows = conn.execute(
            "SELECT usado, COUNT(*) as total FROM sobras WHERE usado != 'NÃO' GROUP BY usado ORDER BY total DESC"
        ).fetchall()

    economia = sum(
        (float(r["largura"]) * float(r["altura"])) / 1_000_000 * 100
        for r in usados_rows
        if r["largura"] and r["altura"]
    )

    return render_template("dashboard.html",
        usados=usados, disponiveis=disponiveis,
        economia=round(economia, 2),
        porcentagem=round(usados / total * 100, 1) if total > 0 else 0,
        ranking={r["usado"]: r["total"] for r in ranking_rows},
        user=session["user"])

# =========================
# LOGOUT
# =========================
@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# =========================
# RUN
# =========================
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)